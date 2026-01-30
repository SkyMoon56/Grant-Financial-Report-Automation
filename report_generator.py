import pandas as pd
import os
import warnings
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- STEP 1: FIND THE FILES ---
possible_paths = [
    os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "SBSC"),
    os.path.join(os.path.expanduser("~"), "Desktop", "SBSC"),
    r"C:\Users\user\OneDrive\Desktop\SBSC"
]

SCRIPT_DIR = None
for p in possible_paths:
    if os.path.exists(p):
        SCRIPT_DIR = p
        break

if not SCRIPT_DIR:
    print("CRITICAL ERROR: Could not find the SBSC folder anywhere.")
    sys.exit()

os.chdir(SCRIPT_DIR)
print(f"Confirmed Working Directory: {SCRIPT_DIR}")

all_files = os.listdir(SCRIPT_DIR)
budget_match = [f for f in all_files if "REFS_BUD" in f]
project_match = [f for f in all_files if "OU_SPNSR" in f]

if not budget_match or not project_match:
    print(f"ERROR: Found folder, but files are missing.")
    print(f"Files inside SBSC: {all_files}")
    sys.exit()

BUDGET_FILE = budget_match[0]
PROJECT_INFO_FILE = project_match[0]
OUTPUT_FILE = os.path.join(os.path.expanduser("~"), "Downloads", "Final_Report.xlsx")

# --- STEP 2: LOAD AND MERGE ---
def load_and_merge_data():
    print(f"Loading {BUDGET_FILE}...")
    main_df = pd.read_excel(BUDGET_FILE, header=1)
    
    print(f"Loading {PROJECT_INFO_FILE}...")
    proj_df = pd.read_excel(PROJECT_INFO_FILE, header=1)

    main_df.columns = main_df.columns.astype(str).str.strip()
    proj_df.columns = proj_df.columns.astype(str).str.strip()

    for df in [main_df, proj_df]:
        for col in ['Project', 'Parent', 'Child']:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

    if 'Child' in proj_df.columns and 'Project' not in proj_df.columns:
        proj_df = proj_df.rename(columns={'Child': 'Project'})

    proj_df = proj_df.rename(columns={'Title': 'Grant Title', 'Sponsor': 'Sponsor Name'})
    merged_df = pd.merge(main_df, proj_df.drop_duplicates(subset=['Project']), on='Project', how='left')

    if 'Parent' in merged_df.columns:
        merged_df['Group_ID'] = merged_df['Parent'].replace(['nan', 'None', ''], None).fillna(merged_df['Project'])
        merged_df = merged_df.sort_values(by=['Group_ID', 'Project'])
    
    if 'Sponsor Name' in merged_df.columns:
        merged_df['Sponsor'] = merged_df['Sponsor Name']

    return merged_df

# --- STEP 3: FORMAT AND EXPORT ---
def format_excel_report(df):
    final_columns = [
        "Budget Type", "Project", "Fund", "Org", "Sponsor",
        "Proj Start Date", "Proj End Date", "PI Name", "Grant Title",
        "Function", "Account", "Budget Amt", "Pre-Encumbered Amt",
        "Encumbered Amt", "Expended Amt", "Remaining Amt"
    ]

    # Create the organized dataframe for export
    result_df = pd.DataFrame()
    for col in final_columns:
        result_df[col] = df[col] if col in df.columns else ""
    
    # Track grouping status (is this row a child?)
    is_child = (df['Project'] != df['Group_ID']).tolist() if 'Group_ID' in df else [False]*len(df)

    print(f"--- Step 3: Formatting {len(result_df)} rows ---")
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Summary"

    # Load rows into Excel
    for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Header Styling
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                # Grouping/Outline Level
                if is_child[r_idx - 2]:
                    ws.row_dimensions[r_idx].outlineLevel = 1
                else:
                    cell.font = Font(bold=True) # Parents are Bold

                # Date Formatting (Cols F & G)
                if c_idx in [6, 7]:
                    cell.number_format = 'mm/dd/yyyy'
                
                # Currency Formatting (Cols L thru P)
                elif c_idx >= 12: 
                    cell.number_format = '"$"#,##0.00'

                # Zebra Striping
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # AUTO-FIT COLUMNS (Fixes the ####### issue)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    # Freeze top row and group settings
    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = False
    
    try:
        wb.save(OUTPUT_FILE)
        print(f"--- SUCCESS --- Saved to Downloads: {OUTPUT_FILE}")
        os.startfile(OUTPUT_FILE)
    except Exception as e:
        print(f"Error saving formatted file: {e}")

if __name__ == "__main__":
    combined_data = load_and_merge_data()
    if not combined_data.empty:
        format_excel_report(combined_data)
    else:
        print("FAILED: Merged data is empty.")
